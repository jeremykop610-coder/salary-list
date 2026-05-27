from __future__ import annotations

import io
import json
import os
import re
import time
import uuid
import zipfile
from http import HTTPStatus
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from urllib.parse import quote, unquote

from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfgen import canvas


BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "outputs"
LOGO_PATH = BASE_DIR / "static" / "andrea-tang-logo.jpg"
OUTPUT_DIR.mkdir(exist_ok=True)
OUTPUT_TTL_SECONDS = 60 * 60

pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))


FIELD_SPECS = [
    ("基本工资", "基本工资", "basic_salary"),
    ("岗位工资", "岗位工资", "position_salary"),
    ("绩效工资", "绩效工资", "performance_salary"),
    ("灵活涨薪", "灵活涨薪", "flex_raise"),
    ("加班费", "加班费", "overtime"),
    ("奖金", "奖金", "bonus"),
    ("请假扣款", "事、病假扣款", "leave_deduction"),
    ("提成", "提成", "commission"),
    ("餐补", "餐补", "meal_allowance"),
    ("其他", "其他", "other"),
    ("小 计", "小计", "subtotal"),
    ("养老保险", "减：养老保险", "pension"),
    ("医疗保险", "减：医疗保险", "medical"),
    ("失业保险", "减：失业保险", "unemployment"),
    ("公 积 金", "减：住房公积金", "housing_fund"),
    ("税前收入", "税前收入", "pre_tax_income"),
    ("个 税", "应纳税额", "tax"),
    ("税后收入", "税后收入", "after_tax_income"),
]


PAYSLIP_FIELD_SPECS = [
    ("基本工资", "basic_salary"),
    ("岗位工资", "position_salary"),
    ("绩效工资", "performance_salary"),
    ("请假扣款", "leave_deduction"),
    ("小    计", "subtotal"),
    ("养老保险", "pension"),
    ("医疗保险", "medical"),
    ("失业保险", "unemployment"),
    ("公 积 金", "housing_fund"),
    ("税前收入", "pre_tax_income"),
    ("个    税", "tax"),
    ("税后收入", "after_tax_income"),
]


def normalize_label(value: Any) -> str:
    return re.sub(r"[\s:：]+", "", str(value or ""))


def build_header_map(ws) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for row_index in range(1, min(ws.max_row, 8) + 1):
        for cell in ws[row_index]:
            label = normalize_label(cell.value)
            if label:
                header_map.setdefault(label, cell.column)
    return header_map


def find_column(header_map: dict[str, int], *names: str) -> int | None:
    normalized = [normalize_label(name) for name in names]
    for name in normalized:
        if name in header_map:
            return header_map[name]
    return None


def amount(value: Any) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, AttributeError):
        return Decimal("0")


def money(value: Any) -> str:
    return f"{amount(value).quantize(Decimal('0.01'))} 元"


def money_number(value: Any) -> str:
    return str(amount(value).quantize(Decimal("0.01")))


def display_month(month: Any) -> str:
    text = str(month or "").strip()
    match = re.search(r"(\d{4})\s*年\s*(\d{1,2})\s*月", text)
    if match:
        return f"{match.group(1)} 年 {match.group(2)} 月"
    return text


def clean_filename(value: str) -> str:
    return re.sub(r'[\\/:*?"<>|\r\n]+', "_", value).strip() or "未命名"


def month_prefix(month: Any) -> str:
    text = str(month or "").strip()
    match = re.search(r"(\d{1,2})\s*月", text)
    return f"{match.group(1)}月" if match else (text or "未知月份")


def make_unique_filename(base: str, used: dict[str, int]) -> str:
    safe_base = clean_filename(base)
    count = used.get(safe_base, 0)
    used[safe_base] = count + 1
    suffix = "" if count == 0 else str(count)
    return f"{safe_base}{suffix}.pdf"


def draw_payslip(row: dict[str, Any]) -> bytes:
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    c.setTitle(f"{month_prefix(row['month'])}工资单-{row['name']}")

    if LOGO_PATH.exists():
        c.drawImage(str(LOGO_PATH), 126.72, height - 121.8 - 42.96, width=345.84, height=42.96)

    c.setFillGray(0.35)
    c.setFont("STSong-Light", 36)
    c.drawCentredString(width / 2, height - 208.01, "工资单")

    c.setFillGray(0.45)
    c.setFont("STSong-Light", 16)
    c.drawCentredString(width / 2, height - 249.53, display_month(row["month"]))

    label_x = 216
    value_x = 300
    y = height - 303.53
    line_height = 27

    c.setFillGray(0.45)
    c.setFont("STSong-Light", 16)
    c.drawString(label_x, y, "姓    名:")
    c.drawString(value_x, y, str(row["name"]))
    y -= line_height

    for index, (display_name, key) in enumerate(PAYSLIP_FIELD_SPECS):
        if index in {5, 11}:
            y -= line_height
        if key == "after_tax_income":
            c.setFillGray(0)
        c.drawString(label_x, y, f"{display_name}:")
        number = money_number(row.get(key))
        c.drawString(value_x, y, number)
        c.drawString(value_x + pdfmetrics.stringWidth(number, "STSong-Light", 16) + 8, y, "元")
        y -= line_height

    c.showPage()
    c.save()
    return buffer.getvalue()


def parse_workbook(file_stream) -> tuple[list[dict[str, Any]], list[str], list[str]]:
    wb = load_workbook(file_stream, data_only=True)
    ws = wb.active
    errors: list[str] = []
    warnings: list[str] = []

    header_map = build_header_map(ws)
    name_col = find_column(header_map, "姓名")
    month_col = find_column(header_map, "月份")

    field_columns: dict[str, int | None] = {}
    for _display, excel_name, key in FIELD_SPECS:
        field_columns[key] = find_column(header_map, excel_name)

    missing_columns = []
    if not name_col:
        missing_columns.append("姓名")
    if not month_col:
        missing_columns.append("月份")
    for display, excel_name, key in FIELD_SPECS:
        if not field_columns[key]:
            missing_columns.append(excel_name or display)

    if missing_columns:
        errors.append("工资表缺少必要列：" + "、".join(missing_columns))
        return [], warnings, errors

    employees: list[dict[str, Any]] = []
    for row_index in range(5, ws.max_row + 1):
        first_value = ws.cell(row_index, 1).value
        if normalize_label(first_value) == "合计":
            break

        name = str(ws.cell(row_index, name_col).value or "").strip()
        month = ws.cell(row_index, month_col).value
        row_has_content = any(ws.cell(row_index, col).value not in (None, "") for col in range(1, ws.max_column + 1))
        if not row_has_content:
            continue

        if not name:
            warnings.append(f"第 {row_index} 行缺少姓名，已跳过。")
            continue
        if not month:
            warnings.append(f"第 {row_index} 行缺少月份，已跳过。")
            continue

        employee = {"name": name, "month": month}
        for _display, _excel_name, key in FIELD_SPECS:
            employee[key] = ws.cell(row_index, field_columns[key]).value
        employees.append(employee)

    if not employees:
        errors.append("没有找到可生成工资单的员工行。请确认第 5 行开始有姓名和月份。")

    return employees, warnings, errors


def json_bytes(payload: dict[str, Any]) -> bytes:
    return json.dumps(payload, ensure_ascii=False).encode("utf-8")


def parse_multipart_file(body: bytes, content_type: str) -> tuple[str, bytes] | None:
    match = re.search(r"boundary=(?P<boundary>[^;]+)", content_type)
    if not match:
        return None
    boundary = match.group("boundary").strip('"').encode()
    marker = b"--" + boundary

    for part in body.split(marker):
        if b"Content-Disposition" not in part:
            continue
        header_blob, _, content = part.partition(b"\r\n\r\n")
        if not content:
            continue
        headers = header_blob.decode("utf-8", errors="ignore")
        if 'name="file"' not in headers:
            continue
        file_match = re.search(r'filename="(?P<filename>[^"]*)"', headers)
        filename = file_match.group("filename") if file_match else "upload.xlsx"
        if content.endswith(b"\r\n"):
            content = content[:-2]
        if content.endswith(b"--"):
            content = content[:-2]
        if content.endswith(b"\r\n"):
            content = content[:-2]
        return filename, content
    return None


def generate_zip(file_name: str, content: bytes) -> tuple[int, dict[str, Any]]:
    cleanup_old_outputs()

    if not file_name.lower().endswith((".xlsx", ".xlsm")):
        return 400, {"ok": False, "errors": ["请上传 .xlsx 或 .xlsm 格式的工资表。"], "warnings": []}

    try:
        employees, warnings, errors = parse_workbook(io.BytesIO(content))
    except Exception as exc:
        return 400, {"ok": False, "errors": [f"无法读取工资表：{exc}"], "warnings": []}

    if errors:
        return 400, {"ok": False, "errors": errors, "warnings": warnings}

    first_month = month_prefix(employees[0]["month"])
    zip_name = clean_filename(f"{first_month}工资单.zip")
    zip_path = OUTPUT_DIR / f"{uuid.uuid4().hex}-{zip_name}"
    used_names: dict[str, int] = {}

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for employee in employees:
            pdf_name = make_unique_filename(f"{month_prefix(employee['month'])}工资单-{employee['name']}", used_names)
            zip_file.writestr(pdf_name, draw_payslip(employee))

        if warnings:
            zip_file.writestr("生成提示.txt", "\n".join(warnings).encode("utf-8"))

    return 200, {
        "ok": True,
        "count": len(employees),
        "warnings": warnings,
        "errors": [],
        "fileName": zip_name,
        "downloadUrl": f"/api/download/{zip_path.name}",
    }


def cleanup_old_outputs() -> None:
    now = time.time()
    for path in OUTPUT_DIR.glob("*.zip"):
        try:
            if now - path.stat().st_mtime > OUTPUT_TTL_SECONDS:
                path.unlink()
        except OSError:
            pass


class PayslipHandler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(BASE_DIR / "static"), **kwargs)

    def send_json(self, status: int, payload: dict[str, Any]) -> None:
        data = json_bytes(payload)
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def do_POST(self) -> None:
        if self.path != "/api/generate":
            self.send_error(HTTPStatus.NOT_FOUND)
            return

        content_length = int(self.headers.get("Content-Length", "0"))
        content_type = self.headers.get("Content-Type", "")
        uploaded = parse_multipart_file(self.rfile.read(content_length), content_type)
        if not uploaded:
            self.send_json(400, {"ok": False, "errors": ["请先选择工资表 Excel 文件。"], "warnings": []})
            return

        status, payload = generate_zip(uploaded[0], uploaded[1])
        self.send_json(status, payload)

    def do_GET(self) -> None:
        if self.path.startswith("/api/download/"):
            file_name = Path(unquote(self.path.removeprefix("/api/download/"))).name
            path = OUTPUT_DIR / file_name
            if not path.exists():
                self.send_json(404, {"ok": False, "errors": ["下载文件不存在或已被清理。"]})
                return

            visible_name = file_name.split("-", 1)[-1]
            data = path.read_bytes()
            self.send_response(200)
            self.send_header("Content-Type", "application/zip")
            self.send_header("Content-Length", str(len(data)))
            self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{quote(visible_name)}")
            self.end_headers()
            self.wfile.write(data)
            return

        if self.path == "/":
            self.path = "/index.html"
        super().do_GET()


if __name__ == "__main__":
    host = os.environ.get("HOST", "0.0.0.0")
    port = int(os.environ.get("PORT", "5001"))
    server = ThreadingHTTPServer((host, port), PayslipHandler)
    print(f"工资单批量生成已启动：http://{host}:{port}")
    server.serve_forever()
